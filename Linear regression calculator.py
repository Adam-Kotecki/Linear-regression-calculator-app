import tkinter as tk
from tkinter import ttk
from tkinter import Entry
from tkinter.messagebox import showinfo
from tkinter import filedialog
import numpy as np
import openpyxl
from sklearn.linear_model import LinearRegression
from matplotlib import pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)

root = tk.Tk()

root.geometry('400x500')
root.resizable(False, False)
root.title('Linear regression calculator')


def main_view():
    root.columnconfigure(0, weight=1)
    root.columnconfigure(1, weight=1)
    
    root.rowconfigure(0, weight=1)
    root.rowconfigure(1, weight=1)
    root.rowconfigure(2, weight=1)
    root.rowconfigure(3, weight=1)
    root.rowconfigure(4, weight=1)
    root.rowconfigure(5, weight=1)
    root.rowconfigure(6, weight=4)
    root.rowconfigure(7, weight=1)
    
    frame = ttk.Frame(root)
    frame.columnconfigure(0, weight=1)
    frame.columnconfigure(1, weight=1)
    
    label1 = tk.Label(root, text="Linear regression")
    label1.grid(column=0, row = 0, columnspan = 4)
    
    label1 = tk.Label(root, text="Please provide numeric values:")
    label1.grid(column=0, row = 1, columnspan = 2)
    
    x_label = tk.Label(root, text="X")
    x_label.grid(column=0, row = 2)
    y_label = tk.Label(root, text="Y")
    y_label.grid(column=1, row = 2)
    
    r = 0
    for i in range(10):
        frame.rowconfigure(r, weight=1)
        globals()[f"x{i + 1}"] = ttk.Entry(frame)
        globals()[f"x{i + 1}"].grid(row = r, column = 0)
        globals()[f"y{i + 1}"] = ttk.Entry(frame)
        globals()[f"y{i + 1}"].grid(row = r, column = 1)
        r += 1
    
    frame.grid(column = 0, row = 3, columnspan = 2, sticky = "N")
    
    global more
    more_btn = tk.Button(root, text="I have more values", command=lambda: more())
    more_btn.grid(column=0, row = 4, columnspan = 2, sticky = "N")
    
    global calculate
    calculate_btn = tk.Button(root, text="Calculate", command=lambda: get_data_from_widgets())
    calculate_btn.grid(column=0, row = 5, columnspan = 2)
    
    global learn
    learn_btn = tk.Button(root, text="Learn about linear regression", command=lambda: learn())
    learn_btn.grid(column=0, row = 6, columnspan = 2)
    
    credit_label = tk.Label(root, text="Created by Adam Kotecki", font = ('TkDefaultFont', 8))
    credit_label.grid(column=0, row = 7, columnspan = 2)

main_view()

def get_data_from_widgets():
    x_list = []
    y_list = []
    for i in range(10):
        if (globals()[f"x{i + 1}"].get() != "") and (globals()[f"y{i + 1}"].get() != ""):
            try:
                float(globals()[f"x{i + 1}"].get())
                float(globals()[f"y{i + 1}"].get())
                x_list.append(float(globals()[f"x{i + 1}"].get()))
                y_list.append(float(globals()[f"y{i + 1}"].get()))
            except:
                tk.messagebox.showinfo(message="All provided values should be numeric.")
                return
    
    if len(x_list) > 1:
        print(x_list)
        print(y_list)
        x_list = np.array(x_list).reshape((-1, 1))
        y_list = np.array(y_list)
        calculate(x_list, y_list)
    else:
        tk.messagebox.showinfo(title = "X, Y values are missing", message="Please provide at least two pairs of values.")


def more():
    root2 = tk.Tk()
    root2.geometry('300x200')
    root2.resizable(False, False)
    root2.title('Get data from Excel')
    root2.rowconfigure(0, weight=2)
    root2.rowconfigure(1, weight=1)
    info = tk.Label(root2, text="Please select file with format xlsx or xlsm. X values should be in the first column, and Y values in the second column. Do not use column headers. All data should be numeric", wraplength = 300, justify = tk.CENTER)
    info.grid(column=0, row = 0)
    select_file = tk.Button(root2, text="Select Excel file", command=lambda: [get_data_from_xlsx(), root2.destroy()])
    select_file.grid(column=0, row = 1)
    root2.mainloop()


def get_data_from_xlsx():
    path = filedialog.askopenfilename()
    try:
        workbook = openpyxl.load_workbook(path)
    except:
       tk.messagebox.showinfo(message="File with wrong format selected. Please select xlsx or xlsm.")
       return
    
    x_list = []
    y_list = []
    sheet = workbook.active
    
    try:
        r = 1
        for cell in sheet['A']:
            if (cell.value != None) and sheet.cell(row=r, column=2).value != None:
                x_list.append(cell.value)
                y_list.append(sheet.cell(row=r, column=2).value)
                r += 1
            else:
                break
    except:
        tk.messagebox.showinfo(message="Something went wrong")
        return
    
    x_list = np.array(x_list).reshape((-1, 1))
    y_list = np.array(y_list)
    calculate(x_list, y_list)
    
    
def calculate(x, y):
    model = LinearRegression()
    model.fit(x, y)
    r_sq = round(model.score(x, y), 4)
    b0 = round(model.intercept_, 4)
    b1 = round(model.coef_[0], 4)
    y_pred = b0 + b1 * x
    
    for widgets in root.winfo_children():
      widgets.destroy()
      
    root.rowconfigure(0, weight=1)
    root.rowconfigure(1, weight=5)
    root.rowconfigure(2, weight=1)
    root.rowconfigure(3, weight=1)
    root.rowconfigure(4, weight=1)
    root.rowconfigure(5, weight=1)
    root.rowconfigure(6, weight=1)
    
    label_title = tk.Label(root, text="Calculated results:")
    label_title.grid(row = 0)
    
    fig = Figure(figsize = (6, 4))
    a = fig.add_subplot(111)
    a.scatter(x, y) 
    a.plot(x, y_pred, color = 'red')
    a.title.set_text("Linear regression")
    canvas = FigureCanvasTkAgg(fig, master = root)  
    canvas.get_tk_widget().grid(row = 1)
    canvas.draw()
    
    text_R2 = tk.Text(root, height = 1, bg = "SystemButtonFace")
    text_R2.insert(1.0, " Coefficient of determination: " + str(r_sq))
    text_R2.configure(state = "disabled")
    text_R2.grid(row = 2)
    
    text_b0 = tk.Text(root, height = 1, bg = "SystemButtonFace")
    text_b0.insert(1.0, " Intercept: " + str(b0))
    text_b0.configure(state = "disabled")
    text_b0.grid(row = 3)
    
    text_b1 = tk.Text(root, height = 1, bg = "SystemButtonFace")
    text_b1.insert(1.0, " Slope: " + str(b1))
    text_b1.configure(state = "disabled")
    text_b1.grid(row = 4)
    
    if b0 > 0:
        equation = "Y = " + str(b1) + " * X " + "+ " + str(b0)
    else:
        equation = "Y = " + str(b1) + " * X " + str(b0)
    
    text_eq = tk.Text(root, height = 1, bg = "SystemButtonFace")
    text_eq.insert(1.0, " Equation: " + equation)
    text_eq.configure(state = "disabled")
    text_eq.grid(row = 5)
    
    back_btn = tk.Button(root, text="Go back", command=lambda: back())
    back_btn.grid(row = 7)
    
def back():
    for widgets in root.winfo_children():
        widgets.destroy()
    main_view()
    
    
def learn():
    for widgets in root.winfo_children():
            widgets.destroy()
            
    root.rowconfigure(0, weight=1)
    root.rowconfigure(1, weight=1)
    root.rowconfigure(2, weight=1)
    root.rowconfigure(3, weight=1)
    root.rowconfigure(4, weight=1)
    root.rowconfigure(5, weight=1)
    root.rowconfigure(6, weight=1)
    
    message1 = tk.Message(width = 350, text = "About linear regression")
    message1.grid(row = 1, columnspan = 2)
    
    message2 = tk.Message(width = 350, text = "Linear regression is a machine learning algorithm that provides a linear relationship between an independent variable (x) and a dependent variable (y) to predict the outputs of future inputs.\n\nIt is also used in data science to determine 'if' and 'to what extent' some phenomenon influences the other. For example, you can examine the dependence of housing prices on distance to the city center. In this case, prices are responses and distances are regressors.")
    message2.grid(row = 2, columnspan = 2)
    
    message4 = tk.Message(width = 350, text = "Model takes the following form:")
    message4.grid(row = 3, columnspan = 2)
    
    message4 = tk.Message(width = 350, text = "y = 𝛽₀ + 𝛽₁𝑥₁ + ⋯ + 𝛽ᵣ𝑥ᵣ + e", font = ("TkDefaultFont", 12))
    message4.grid(row = 4, columnspan = 2)
    
    message6 = tk.Message(width = 350, text = "Where:\n\n𝛽₀ - Intercept - the mean value of the response variable when x = 0. In some cases, it makes sense to interpret this value but not always.\n\n𝛽₁ - Slope - the average change in the response variable for a one unit increase in x.\n\ne is the random error.")
    message6.grid(row = 5, columnspan = 2)
    
    message5 = tk.Message(width = 350, text = "Once you've built your linear regression model, you can evaluate it using the coefficient of determination (R²), which measures how well a model predicts an outcome. R-squared values range from 0 to 1. The better a model is, the closer this measure will be to 1.")
    message5.grid(row = 6, columnspan = 2)
    
    back_btn = tk.Button(root, text="Go back", command=lambda: back())
    back_btn.grid(row = 7, columnspan = 2)
    
    
root.mainloop()